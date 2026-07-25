#!/usr/bin/env python3
"""
XLSX文档转换器
将XLSX文档转换为HTML格式，优化移动端显示
"""

import logging
from typing import Tuple, Optional

from .base import BaseConverter, HTMLWrapper, ConversionError

logger = logging.getLogger(__name__)


class XlsxConverter(BaseConverter):
    """XLSX文档转换器"""
    
    def __init__(self):
        super().__init__()
        self.supported_formats = ['xlsx']
    
    def is_supported(self, file_type: str) -> bool:
        """检查文件类型是否支持转换"""
        return file_type.lower() in self.supported_formats
    
    def convert_to_html(self, file_path: str, file_type: str) -> Tuple[bool, Optional[str], Optional[str]]:
        """转换XLSX文档为HTML，优化移动端显示"""
        try:
            # 验证文件
            is_valid, error_msg = self.validate_file(file_path)
            if not is_valid:
                return False, None, error_msg
            
            from openpyxl import load_workbook

            workbook = load_workbook(file_path, read_only=True)
            content_parts = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # 获取有数据的区域
                max_row = min(sheet.max_row, 100)  # 限制最多100行
                max_col = min(sheet.max_column, 15)  # 移动端限制为15列

                if max_row == 0 or max_col == 0:
                    content_parts.append(f"<h3>工作表: {HTMLWrapper.escape_html(sheet_name)} (空表)</h3>")
                    continue

                content_parts.append(f"<div class='excel-sheet'>")
                content_parts.append(f"<h3 class='sheet-title'>📊 {HTMLWrapper.escape_html(sheet_name)}</h3>")

                # 添加表格信息
                actual_rows = sheet.max_row
                actual_cols = sheet.max_column
                content_parts.append(f"<div class='sheet-info'>")
                content_parts.append(f"<span class='info-item'>📏 {actual_rows} 行 × {actual_cols} 列</span>")
                if actual_rows > 100:
                    content_parts.append(f"<span class='info-item'>⚠️ 仅显示前100行</span>")
                if actual_cols > 15:
                    content_parts.append(f"<span class='info-item'>⚠️ 仅显示前15列</span>")
                content_parts.append(f"</div>")

                # 创建移动端优化的表格
                content_parts.append("<div class='excel-table-wrapper'>")
                content_parts.append("<table class='excel-table'>")

                # 添加列标题（A, B, C...）
                content_parts.append("<thead><tr><th class='row-header'>#</th>")
                for col_idx in range(max_col):
                    col_letter = HTMLWrapper.get_column_letter(col_idx + 1)
                    content_parts.append(f"<th class='col-header'>{col_letter}</th>")
                content_parts.append("</tr></thead>")

                content_parts.append("<tbody>")
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), 1):
                    content_parts.append("<tr>")
                    # 行号
                    content_parts.append(f"<td class='row-header'>{row_idx}</td>")

                    for cell in row:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        # 截断过长的内容
                        if len(cell_value) > 50:
                            display_value = cell_value[:47] + "..."
                            content_parts.append(f"<td class='excel-cell' title='{HTMLWrapper.escape_html(cell_value)}'>{HTMLWrapper.escape_html(display_value)}</td>")
                        else:
                            content_parts.append(f"<td class='excel-cell'>{HTMLWrapper.escape_html(cell_value)}</td>")
                    content_parts.append("</tr>")

                content_parts.append("</tbody></table>")
                content_parts.append("</div>")  # excel-table-wrapper
                content_parts.append("</div>")  # excel-sheet

            workbook.close()
            html_content = "\n".join(content_parts)
            full_html = self._wrap_excel_html(html_content, "Excel文档预览")

            return True, full_html, None

        except Exception as e:
            logger.error(f"转换XLSX失败: {str(e)}")
            return False, None, str(e)
    
    def _wrap_excel_html(self, content: str, title: str) -> str:
        """专门为Excel文档包装HTML，优化移动端表格显示"""
        escaped_title = HTMLWrapper.escape_html(title)
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0, user-scalable=yes, minimum-scale=0.5, maximum-scale=3.0">
            <meta name="format-detection" content="telephone=no">
            <meta name="apple-mobile-web-app-capable" content="yes">
            <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
            <title>{escaped_title}</title>
            <style>
                * {{
                    box-sizing: border-box;
                }}
                body {{
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
                    line-height: 1.4;
                    margin: 0;
                    padding: 5px;
                    background: #f8f9fa;
                    color: #333;
                    font-size: 13px;
                    -webkit-text-size-adjust: 100%;
                    -webkit-font-smoothing: antialiased;
                }}
                
                .excel-sheet {{
                    background: white;
                    margin-bottom: 15px;
                    border-radius: 6px;
                    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
                    overflow: hidden;
                }}
                
                .sheet-title {{
                    background: linear-gradient(135deg, #007bff, #0056b3);
                    color: white;
                    margin: 0;
                    padding: 10px 15px;
                    font-size: 16px;
                    font-weight: 600;
                    text-shadow: 0 1px 2px rgba(0,0,0,0.1);
                }}
                
                .sheet-info {{
                    background: #e9ecef;
                    padding: 8px 15px;
                    border-bottom: 1px solid #dee2e6;
                    font-size: 12px;
                    color: #495057;
                }}
                
                .info-item {{
                    display: inline-block;
                    margin-right: 15px;
                    padding: 2px 6px;
                    background: white;
                    border-radius: 3px;
                    border: 1px solid #ced4da;
                }}
                
                .excel-table-wrapper {{
                    overflow-x: auto;
                    -webkit-overflow-scrolling: touch;
                    max-height: 70vh;
                    overflow-y: auto;
                }}
                
                .excel-table {{
                    width: 100%;
                    border-collapse: collapse;
                    font-size: 12px;
                    min-width: 600px;
                }}
                
                .excel-table th,
                .excel-table td {{
                    border: 1px solid #dee2e6;
                    padding: 6px 8px;
                    text-align: left;
                    vertical-align: top;
                    white-space: nowrap;
                    overflow: hidden;
                    text-overflow: ellipsis;
                    max-width: 150px;
                }}
                
                .excel-table th {{
                    background: #f8f9fa;
                    font-weight: 600;
                    color: #495057;
                    position: sticky;
                    top: 0;
                    z-index: 10;
                }}
                
                .row-header {{
                    background: #e9ecef !important;
                    font-weight: 600;
                    text-align: center;
                    min-width: 40px;
                    max-width: 40px;
                    position: sticky;
                    left: 0;
                    z-index: 5;
                }}
                
                .col-header {{
                    background: #f1f3f4 !important;
                    text-align: center;
                    min-width: 80px;
                }}
                
                .excel-cell {{
                    background: white;
                    transition: background-color 0.2s;
                }}
                
                .excel-cell:hover {{
                    background: #f8f9fa;
                }}
                
                .excel-table tbody tr:nth-child(even) .excel-cell {{
                    background: #fdfdfe;
                }}
                
                .excel-table tbody tr:nth-child(even) .excel-cell:hover {{
                    background: #f8f9fa;
                }}
                
                /* 移动端优化 */
                @media (max-width: 768px) {{
                    body {{
                        padding: 2px;
                        font-size: 12px;
                    }}
                    
                    .sheet-title {{
                        padding: 8px 10px;
                        font-size: 14px;
                    }}
                    
                    .sheet-info {{
                        padding: 6px 10px;
                        font-size: 11px;
                    }}
                    
                    .info-item {{
                        margin-right: 8px;
                        padding: 1px 4px;
                        font-size: 10px;
                    }}
                    
                    .excel-table {{
                        font-size: 11px;
                        min-width: 500px;
                    }}
                    
                    .excel-table th,
                    .excel-table td {{
                        padding: 4px 6px;
                        max-width: 100px;
                    }}
                    
                    .row-header {{
                        min-width: 30px;
                        max-width: 30px;
                        font-size: 10px;
                    }}
                    
                    .col-header {{
                        min-width: 60px;
                        font-size: 10px;
                    }}
                }}
                
                /* 滚动条样式 */
                .excel-table-wrapper::-webkit-scrollbar {{
                    width: 8px;
                    height: 8px;
                }}
                
                .excel-table-wrapper::-webkit-scrollbar-track {{
                    background: #f1f3f4;
                    border-radius: 4px;
                }}
                
                .excel-table-wrapper::-webkit-scrollbar-thumb {{
                    background: #c1c8cd;
                    border-radius: 4px;
                }}
                
                .excel-table-wrapper::-webkit-scrollbar-thumb:hover {{
                    background: #9aa0a6;
                }}
            </style>
        </head>
        <body>
            {content}
            
            <script>
                document.addEventListener('DOMContentLoaded', function() {{
                    console.log('Excel预览页面加载完成');
                    
                    // 优化表格滚动体验
                    const wrappers = document.querySelectorAll('.excel-table-wrapper');
                    wrappers.forEach(wrapper => {{
                        // 添加滚动指示器
                        wrapper.addEventListener('scroll', function() {{
                            const scrollLeft = this.scrollLeft;
                            const scrollTop = this.scrollTop;
                            
                            // 可以在这里添加滚动位置指示器
                            console.log('表格滚动位置:', {{ left: scrollLeft, top: scrollTop }});
                        }});
                    }});
                }});
            </script>
        </body>
        </html>
        """
